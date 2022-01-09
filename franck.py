
from openpyxl import load_workbook
#load in the workbook
wb = load_workbook("Employeedata.xlsx")
sheet = wb.active 
for i in range(2,sheet.max_row +1): 
    cell = sheet.cell(i,2)  #choosing the column i am going to work with
    if "helpinghands.cm" in cell.value:
        updated_email= (cell.value).replace('helpinghands.cm','handsinhands.org') #replacing the old domaine by the new domaine
        sheet.cell(i,2).value = updated_email
        #now we are creating the new file with the ne domaine name  
wb.save("new.xlsx")       